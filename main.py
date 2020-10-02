import openpyxl
import os
import platform
from multiprocessing import Process
import json

process_number= []
startrow= 0

platform_name = platform.system()

def load_json_file(filename: str) -> dict:
    with open(filename, encoding='utf-8') as file_stream:
        data = json.load(file_stream)
    return data

def fix_json_file(filename: str, filedata: dict):
    with open(filename, 'w', encoding='utf-8') as file_stream:
        json.dump(filedata, file_stream, ensure_ascii=False, indent="\t")

def make_json_file(filename: str):
    try:
        if platform_name == "Windows":
            os.system('del log.json')
        else:
            os.system('rm -rf ./log.json')
    except:
        with open(filename, 'w', encoding='utf-8') as _:
            pass

credentials= load_json_file('credentials.json')

id= credentials['id']
pw= credentials['pw']
list= credentials['list']

try:
    totprocess= int(credentials['multiprocess'])
except ValueError:
    totprocess= 1

for i in range(totprocess):
    process_number.append(0)

wb= openpyxl.load_workbook(list, data_only= True)
ws= wb['Sheet1']

try:
    log= load_json_file('log.json')
    if log['list_name'] == list:
        startrow= int(log['last_row_no'])

except :
    make_json_file('log.json')

    while True:
        try:
            video_id = int(ws.cell(startrow + 1, 2).value)
        except ValueError:
            startrow = startrow + 1
        else:
            break

    json_data= dict()
    json_data["list_name"] = str(list)
    json_data["last_row_no"] = str(startrow)
    json_data["last_vid_id"] = ""

    fix_json_file('log.json', json_data)



def run(url, id, pw):
    os.system(f'youtube-dl --username \"{id}\" --password \"{pw}\" \"{url}\"')

def job(processnum: int):
    global list

    while True:
        row = int(startrow + processnum + process_number[processnum])
        urlid = str(ws.cell(row, 2).value)

        if processnum == 0:

            json_data = dict()
            json_data["list_name"] = str(list)
            json_data["last_row_no"] = str(row)
            json_data["last_vid_id"] = str(urlid)

            fix_json_file('log.json', json_data)

        if urlid.isdigit():
            url = f'https://vlive.tv/video/{urlid}'
            print(f"{url} {row}")
            run(url, id, pw)
        elif urlid == 'None':
            break
        else:
            pass

        process_number[processnum] = process_number[processnum] + totprocess


if __name__ == "__main__":
    procs= []

    for i in range(totprocess):
        proc= Process(target= job, args= (i,),)
        proc.daemon = True
        print(f"========Process {i} is running========")
        proc.start()
        procs.append(proc)

    for proc in procs:
        proc.join()